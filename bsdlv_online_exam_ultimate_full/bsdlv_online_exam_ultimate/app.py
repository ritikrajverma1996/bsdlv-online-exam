
import os, sqlite3, uuid, io, json
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, g
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'instance', 'exam.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key')

TRANSLATIONS = {
    "en": {
        "site_name": "BSDLV Online Exam",
        "admin_login": "Admin Login",
        "student_login": "Student Login",
        "login": "Login",
        "logout": "Logout",
        "dashboard": "Dashboard",
        "students": "Students",
        "questions": "Questions",
        "exams": "Exams",
        "results": "Results",
        "welcome": "Welcome",
        "language": "Language",
        "instructions": "Instructions",
        "start_exam": "Start Exam",
        "submit_exam": "Submit Exam",
        "result_card": "Result Card",
        "admit_card": "Admit Card",
    },
    "hi": {
        "site_name": "BSDLV Online Exam",
        "admin_login": "एडमिन लॉगिन",
        "student_login": "छात्र लॉगिन",
        "login": "लॉगिन",
        "logout": "लॉगआउट",
        "dashboard": "डैशबोर्ड",
        "students": "छात्र",
        "questions": "प्रश्न",
        "exams": "परीक्षाएँ",
        "results": "परिणाम",
        "welcome": "स्वागत है",
        "language": "भाषा",
        "instructions": "निर्देश",
        "start_exam": "परीक्षा शुरू करें",
        "submit_exam": "परीक्षा जमा करें",
        "result_card": "परिणाम पत्र",
        "admit_card": "प्रवेश पत्र",
    }
}

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(error=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def q(sql, params=(), one=False, commit=False):
    db = get_db()
    cur = db.execute(sql, params)
    if commit:
        db.commit()
    if sql.lstrip().upper().startswith("SELECT"):
        rows = cur.fetchall()
        return (rows[0] if rows else None) if one else rows
    return cur.lastrowid

def init_db():
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS admins (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        password TEXT
    );
    CREATE TABLE IF NOT EXISTS students (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT UNIQUE,
        name TEXT,
        student_class TEXT,
        section TEXT,
        password TEXT,
        photo TEXT
    );
    CREATE TABLE IF NOT EXISTS questions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        question TEXT,
        optionA TEXT,
        optionB TEXT,
        optionC TEXT,
        optionD TEXT,
        correctAnswer TEXT,
        marks REAL DEFAULT 1,
        difficulty TEXT,
        subject TEXT,
        chapter TEXT
    );
    CREATE TABLE IF NOT EXISTS exams (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        subject TEXT,
        student_class TEXT,
        section TEXT,
        duration_minutes INTEGER DEFAULT 30,
        negative_marks REAL DEFAULT 0,
        pass_marks REAL DEFAULT 0,
        instructions TEXT,
        question_ids_json TEXT,
        is_active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        exam_id INTEGER,
        student_id INTEGER,
        score REAL,
        total_marks REAL,
        answers_json TEXT,
        started_at TEXT,
        submitted_at TEXT,
        UNIQUE(exam_id, student_id)
    );
    """)
    db.commit()
    admin = cur.execute("SELECT id FROM admins WHERE username='admin'").fetchone()
    if not admin:
        cur.execute("INSERT INTO admins(username,password) VALUES(?,?)", ('admin','admin123'))
    if not cur.execute("SELECT value FROM settings WHERE key='site_title'").fetchone():
        cur.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ('site_title','BSDLV Online Exam'))
    if not cur.execute("SELECT value FROM settings WHERE key='logo'").fetchone():
        cur.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ('logo',''))
    db.commit()
    db.close()

def setting(key, default=''):
    row = q("SELECT value FROM settings WHERE key=?", (key,), one=True)
    return row['value'] if row else default

@app.context_processor
def inject_globals():
    lang = session.get('lang', 'en')
    return {
        'T': TRANSLATIONS.get(lang, TRANSLATIONS['en']),
        'lang': lang,
        'site_title': setting('site_title', 'BSDLV Online Exam'),
        'logo_file': setting('logo', '')
    }

@app.route('/switch-language/<lang_code>')
def switch_language(lang_code):
    if lang_code in ('en', 'hi'):
        session['lang'] = lang_code
    return redirect(request.referrer or url_for('index'))

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get('admin_id'):
            return fn(*args, **kwargs)
        flash('Admin login required')
        return redirect(url_for('admin_login'))
    return wrapper

def student_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if session.get('student_id'):
            return fn(*args, **kwargs)
        flash('Student login required')
        return redirect(url_for('student_login'))
    return wrapper

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        user = q("SELECT * FROM admins WHERE username=? AND password=?", (request.form['username'], request.form['password']), one=True)
        if user:
            session.clear()
            session['admin_id'] = user['id']
            session['lang'] = session.get('lang', 'en')
            return redirect(url_for('admin_dashboard'))
        flash('Invalid admin login')
    return render_template('admin_login.html')

@app.route('/student/login', methods=['GET', 'POST'])
def student_login():
    if request.method == 'POST':
        user = q("SELECT * FROM students WHERE student_id=? AND password=?", (request.form['student_id'], request.form['password']), one=True)
        if user:
            session.clear()
            session['student_id'] = user['id']
            session['lang'] = session.get('lang', 'en')
            return redirect(url_for('student_dashboard'))
        flash('Invalid student login')
    return render_template('student_login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin')
@admin_required
def admin_dashboard():
    stats = {
        'students': q("SELECT COUNT(*) c FROM students", one=True)['c'],
        'questions': q("SELECT COUNT(*) c FROM questions", one=True)['c'],
        'exams': q("SELECT COUNT(*) c FROM exams", one=True)['c'],
        'results': q("SELECT COUNT(*) c FROM results", one=True)['c'],
    }
    return render_template('admin_dashboard.html', stats=stats)

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    if request.method == 'POST':
        title = request.form.get('site_title', 'BSDLV Online Exam').strip() or 'BSDLV Online Exam'
        q("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ('site_title', title), commit=True)
        f = request.files.get('logo')
        if f and f.filename:
            filename = f"logo_{uuid.uuid4().hex}_{secure_filename(f.filename)}"
            f.save(os.path.join(UPLOAD_DIR, filename))
            q("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ('logo', filename), commit=True)
        if request.form.get('new_password'):
            q("UPDATE admins SET password=? WHERE id=?", (request.form['new_password'], session['admin_id']), commit=True)
            flash('Admin password updated')
        flash('Settings saved')
        return redirect(url_for('admin_settings'))
    return render_template('admin_settings.html')

@app.route('/admin/students')
@admin_required
def admin_students():
    rows = q("SELECT * FROM students ORDER BY id DESC")
    return render_template('students.html', students=rows)

@app.route('/admin/students/add', methods=['POST'])
@admin_required
def add_student():
    photo_name = ''
    f = request.files.get('photo')
    if f and f.filename:
        photo_name = f"student_{uuid.uuid4().hex}_{secure_filename(f.filename)}"
        f.save(os.path.join(UPLOAD_DIR, photo_name))
    q("""INSERT INTO students(student_id,name,student_class,section,password,photo)
         VALUES(?,?,?,?,?,?)""",
      (request.form['student_id'], request.form['name'], request.form.get('student_class',''),
       request.form.get('section',''), request.form['password'], photo_name), commit=True)
    flash('Student added')
    return redirect(url_for('admin_students'))

@app.route('/admin/students/import', methods=['POST'])
@admin_required
def import_students():
    f = request.files.get('file')
    if not f:
        flash('No file selected')
        return redirect(url_for('admin_students'))
    wb = load_workbook(f)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    header_map = {h:i for i,h in enumerate(headers)}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vals = {h: row[idx] if idx < len(row) else '' for h, idx in header_map.items()}
        exists = q("SELECT id FROM students WHERE student_id=?", (str(vals.get('student_id','')).strip(),), one=True)
        if exists:
            continue
        q("""INSERT INTO students(student_id,name,student_class,section,password,photo)
             VALUES(?,?,?,?,?,?)""", (
                 str(vals.get('student_id','')).strip(),
                 str(vals.get('name','')).strip(),
                 str(vals.get('student_class','')).strip(),
                 str(vals.get('section','')).strip(),
                 str(vals.get('password','1234')).strip() or '1234',
                 ''
             ), commit=True)
        count += 1
    flash(f'{count} students imported')
    return redirect(url_for('admin_students'))

@app.route('/admin/students/<int:student_id>/delete')
@admin_required
def delete_student(student_id):
    res = q("SELECT id FROM results WHERE student_id=?", (student_id,), one=True)
    if res:
        flash('Cannot delete student with saved result')
    else:
        q("DELETE FROM students WHERE id=?", (student_id,), commit=True)
        flash('Student deleted')
    return redirect(url_for('admin_students'))

@app.route('/admin/questions')
@admin_required
def admin_questions():
    search = request.args.get('search','').strip()
    subject = request.args.get('subject','').strip()
    sql = "SELECT * FROM questions WHERE 1=1"
    params = []
    if search:
        sql += " AND question LIKE ?"
        params.append(f'%{search}%')
    if subject:
        sql += " AND subject LIKE ?"
        params.append(f'%{subject}%')
    sql += " ORDER BY id DESC"
    rows = q(sql, tuple(params))
    return render_template('questions.html', questions=rows, search=search, subject=subject)

@app.route('/admin/questions/add', methods=['POST'])
@admin_required
def add_question():
    q("""INSERT INTO questions(question,optionA,optionB,optionC,optionD,correctAnswer,marks,difficulty,subject,chapter)
         VALUES(?,?,?,?,?,?,?,?,?,?)""", (
             request.form['question'], request.form['optionA'], request.form['optionB'],
             request.form['optionC'], request.form['optionD'], request.form['correctAnswer'],
             float(request.form.get('marks') or 1), request.form.get('difficulty','Medium'),
             request.form.get('subject',''), request.form.get('chapter','')
         ), commit=True)
    flash('Question added')
    return redirect(url_for('admin_questions'))

@app.route('/admin/questions/<int:qid>/edit', methods=['GET','POST'])
@admin_required
def edit_question(qid):
    row = q("SELECT * FROM questions WHERE id=?", (qid,), one=True)
    if not row:
        flash('Question not found')
        return redirect(url_for('admin_questions'))
    if request.method == 'POST':
        q("""UPDATE questions SET question=?, optionA=?, optionB=?, optionC=?, optionD=?, correctAnswer=?, marks=?, difficulty=?, subject=?, chapter=? WHERE id=?""",
          (request.form['question'], request.form['optionA'], request.form['optionB'],
           request.form['optionC'], request.form['optionD'], request.form['correctAnswer'],
           float(request.form.get('marks') or 1), request.form.get('difficulty','Medium'),
           request.form.get('subject',''), request.form.get('chapter',''), qid), commit=True)
        flash('Question updated')
        return redirect(url_for('admin_questions'))
    return render_template('edit_question.html', qrow=row)

@app.route('/admin/questions/<int:qid>/delete')
@admin_required
def delete_question(qid):
    q("DELETE FROM questions WHERE id=?", (qid,), commit=True)
    flash('Question deleted')
    return redirect(url_for('admin_questions'))

@app.route('/admin/questions/import', methods=['POST'])
@admin_required
def import_questions():
    f = request.files.get('file')
    if not f:
        flash('No file selected')
        return redirect(url_for('admin_questions'))
    wb = load_workbook(f)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    header_map = {h:i for i,h in enumerate(headers)}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vals = {h: row[idx] if idx < len(row) else '' for h, idx in header_map.items()}
        question_text = str(vals.get('question','')).strip()
        exists = q("SELECT id FROM questions WHERE question=?", (question_text,), one=True)
        if exists:
            continue
        q("""INSERT INTO questions(question,optionA,optionB,optionC,optionD,correctAnswer,marks,difficulty,subject,chapter)
             VALUES(?,?,?,?,?,?,?,?,?,?)""", (
                 question_text, str(vals.get('optionA','')).strip(), str(vals.get('optionB','')).strip(),
                 str(vals.get('optionC','')).strip(), str(vals.get('optionD','')).strip(),
                 str(vals.get('correctAnswer','A')).strip().upper(), float(vals.get('marks',1) or 1),
                 str(vals.get('difficulty','Medium')).strip(), str(vals.get('subject','')).strip(),
                 str(vals.get('chapter','')).strip()
             ), commit=True)
        count += 1
    flash(f'{count} questions imported')
    return redirect(url_for('admin_questions'))

@app.route('/admin/exams')
@admin_required
def admin_exams():
    exams = q("SELECT * FROM exams ORDER BY id DESC")
    questions = q("SELECT * FROM questions ORDER BY id DESC LIMIT 500")
    return render_template('exams.html', exams=exams, questions=questions)

@app.route('/admin/exams/add', methods=['POST'])
@admin_required
def add_exam():
    selected = request.form.getlist('question_ids')
    if not selected:
        flash('Select at least one question')
        return redirect(url_for('admin_exams'))
    q("""INSERT INTO exams(title,subject,student_class,section,duration_minutes,negative_marks,pass_marks,instructions,question_ids_json,is_active)
         VALUES(?,?,?,?,?,?,?,?,?,1)""",
      (request.form['title'], request.form.get('subject',''), request.form.get('student_class',''),
       request.form.get('section',''), int(request.form.get('duration_minutes') or 30),
       float(request.form.get('negative_marks') or 0), float(request.form.get('pass_marks') or 0),
       request.form.get('instructions',''), json.dumps([int(x) for x in selected])), commit=True)
    flash('Exam created')
    return redirect(url_for('admin_exams'))

@app.route('/admin/exams/<int:eid>/edit', methods=['GET','POST'])
@admin_required
def edit_exam(eid):
    exam = q("SELECT * FROM exams WHERE id=?", (eid,), one=True)
    if not exam:
        flash('Exam not found')
        return redirect(url_for('admin_exams'))
    questions = q("SELECT * FROM questions ORDER BY id DESC LIMIT 500")
    chosen = set(json.loads(exam['question_ids_json'] or '[]'))
    if request.method == 'POST':
        selected = [int(x) for x in request.form.getlist('question_ids')]
        q("""UPDATE exams SET title=?,subject=?,student_class=?,section=?,duration_minutes=?,negative_marks=?,pass_marks=?,instructions=?,question_ids_json=?,is_active=? WHERE id=?""",
          (request.form['title'], request.form.get('subject',''), request.form.get('student_class',''),
           request.form.get('section',''), int(request.form.get('duration_minutes') or 30),
           float(request.form.get('negative_marks') or 0), float(request.form.get('pass_marks') or 0),
           request.form.get('instructions',''), json.dumps(selected), 1 if request.form.get('is_active') else 0, eid), commit=True)
        flash('Exam updated')
        return redirect(url_for('admin_exams'))
    return render_template('edit_exam.html', exam=exam, questions=questions, chosen=chosen)

@app.route('/admin/exams/<int:eid>/delete')
@admin_required
def delete_exam(eid):
    q("DELETE FROM exams WHERE id=?", (eid,), commit=True)
    flash('Exam deleted')
    return redirect(url_for('admin_exams'))

@app.route('/admin/results')
@admin_required
def admin_results():
    rows = q("""SELECT r.*, e.title exam_title, s.name student_name, s.student_id roll_no
                FROM results r
                JOIN exams e ON e.id=r.exam_id
                JOIN students s ON s.id=r.student_id
                ORDER BY r.id DESC""")
    return render_template('results.html', results=rows)

def student_exams_for(student):
    return q("""SELECT * FROM exams
                WHERE is_active=1
                  AND (student_class='' OR student_class=?)
                  AND (section='' OR section=?)
                ORDER BY id DESC""", (student['student_class'], student['section']))

@app.route('/student')
@student_required
def student_dashboard():
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exams = student_exams_for(student)
    done = {r['exam_id']: r for r in q("SELECT * FROM results WHERE student_id=?", (student['id'],))}
    return render_template('student_dashboard.html', student=student, exams=exams, done=done)

@app.route('/student/admit-card')
@student_required
def student_admit_card():
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exams = student_exams_for(student)
    return render_template('admit_card.html', student=student, exams=exams)

@app.route('/student/admit-card/pdf')
@student_required
def student_admit_card_pdf():
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exams = student_exams_for(student)
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    y = 800
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, y, setting('site_title', 'BSDLV Online Exam'))
    y -= 30
    p.setFont("Helvetica", 12)
    p.drawString(50, y, f"Admit Card / प्रवेश पत्र")
    y -= 25
    for line in [f"Name: {student['name']}", f"Student ID: {student['student_id']}",
                 f"Class: {student['student_class']}", f"Section: {student['section']}"]:
        p.drawString(50, y, line); y -= 20
    y -= 10
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Assigned Exams")
    y -= 20
    p.setFont("Helvetica", 11)
    for ex in exams:
        p.drawString(60, y, f"- {ex['title']} ({ex['subject']}) {ex['duration_minutes']} min")
        y -= 18
        if y < 70:
            p.showPage(); y = 800
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="admit_card.pdf", mimetype='application/pdf')

@app.route('/student/exam/<int:eid>/instructions')
@student_required
def exam_instructions(eid):
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exam = q("""SELECT * FROM exams WHERE id=? AND is_active=1
                AND (student_class='' OR student_class=?)
                AND (section='' OR section=?)""", (eid, student['student_class'], student['section']), one=True)
    if not exam:
        flash('Exam not available')
        return redirect(url_for('student_dashboard'))
    already = q("SELECT id FROM results WHERE exam_id=? AND student_id=?", (eid, student['id']), one=True)
    if already:
        flash('You have already attempted this exam')
        return redirect(url_for('student_dashboard'))
    return render_template('exam_instructions.html', exam=exam)

@app.route('/student/exam/<int:eid>/start')
@student_required
def start_exam(eid):
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exam = q("""SELECT * FROM exams WHERE id=? AND is_active=1
                AND (student_class='' OR student_class=?)
                AND (section='' OR section=?)""", (eid, student['student_class'], student['section']), one=True)
    if not exam:
        flash('Exam not available')
        return redirect(url_for('student_dashboard'))
    already = q("SELECT id FROM results WHERE exam_id=? AND student_id=?", (eid, student['id']), one=True)
    if already:
        flash('You have already attempted this exam')
        return redirect(url_for('student_dashboard'))
    ids = json.loads(exam['question_ids_json'] or '[]')
    placeholders = ",".join(["?"]*len(ids)) if ids else "0"
    questions = q(f"SELECT * FROM questions WHERE id IN ({placeholders})", tuple(ids)) if ids else []
    return render_template('take_exam.html', exam=exam, questions=questions, now=datetime.utcnow().isoformat())

@app.route('/student/exam/<int:eid>/submit', methods=['POST'])
@student_required
def submit_exam(eid):
    student = q("SELECT * FROM students WHERE id=?", (session['student_id'],), one=True)
    exam = q("SELECT * FROM exams WHERE id=?", (eid,), one=True)
    if not exam:
        flash('Exam not found')
        return redirect(url_for('student_dashboard'))
    already = q("SELECT id FROM results WHERE exam_id=? AND student_id=?", (eid, student['id']), one=True)
    if already:
        flash('Already submitted')
        return redirect(url_for('student_dashboard'))
    ids = json.loads(exam['question_ids_json'] or '[]')
    score = 0.0
    total = 0.0
    answers = {}
    if ids:
        placeholders = ",".join(["?"]*len(ids))
        questions = q(f"SELECT * FROM questions WHERE id IN ({placeholders})", tuple(ids))
    else:
        questions = []
    for qu in questions:
        selected = request.form.get(f"q_{qu['id']}", '')
        answers[str(qu['id'])] = selected
        total += float(qu['marks'] or 1)
        if selected == qu['correctAnswer']:
            score += float(qu['marks'] or 1)
        elif selected:
            score -= float(exam['negative_marks'] or 0)
    if score < 0:
        score = 0
    q("""INSERT INTO results(exam_id,student_id,score,total_marks,answers_json,started_at,submitted_at)
         VALUES(?,?,?,?,?,?,?)""",
      (eid, student['id'], round(score,2), round(total,2), json.dumps(answers),
       request.form.get('started_at',''), datetime.utcnow().isoformat()), commit=True)
    flash('Exam submitted')
    row = q("SELECT id FROM results WHERE exam_id=? AND student_id=?", (eid, student['id']), one=True)
    return redirect(url_for('student_result', rid=row['id']))

@app.route('/student/result/<int:rid>')
@student_required
def student_result(rid):
    row = q("""SELECT r.*, e.title exam_title, e.pass_marks, s.name student_name, s.student_id roll_no
               FROM results r
               JOIN exams e ON e.id=r.exam_id
               JOIN students s ON s.id=r.student_id
               WHERE r.id=? AND r.student_id=?""", (rid, session['student_id']), one=True)
    if not row:
        flash('Result not found')
        return redirect(url_for('student_dashboard'))
    return render_template('student_result.html', row=row)


def build_result_pdf_response(row):
    return build_result_pdf_response(row)


@app.route('/student/result/<int:rid>/pdf')
@student_required
def student_result_pdf(rid):
    row = q("""SELECT r.*, e.title exam_title, e.pass_marks, s.name student_name, s.student_id roll_no, s.student_class, s.section
               FROM results r
               JOIN exams e ON e.id=r.exam_id
               JOIN students s ON s.id=r.student_id
               WHERE r.id=? AND r.student_id=?""", (rid, session['student_id']), one=True)
    if not row:
        flash('Result not found')
        return redirect(url_for('student_dashboard'))
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    y=800
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50,y, setting('site_title','BSDLV Online Exam')); y-=30
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50,y, "Result Card / परिणाम पत्र"); y-=25
    p.setFont("Helvetica", 12)
    for line in [
        f"Name: {row['student_name']}",
        f"Student ID: {row['roll_no']}",
        f"Class/Section: {row['student_class']} / {row['section']}",
        f"Exam: {row['exam_title']}",
        f"Score: {row['score']} / {row['total_marks']}",
        f"Pass Marks: {row['pass_marks']}",
        f"Status: {'PASS' if row['score'] >= row['pass_marks'] else 'FAIL'}"
    ]:
        p.drawString(50,y,line); y-=20
    p.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="result_card.pdf", mimetype='application/pdf')

@app.route('/admin/result/<int:rid>/pdf')
@admin_required
def admin_result_pdf(rid):
    row = q("""SELECT r.*, e.title exam_title, e.pass_marks, s.name student_name, s.student_id roll_no, s.student_class, s.section
               FROM results r
               JOIN exams e ON e.id=r.exam_id
               JOIN students s ON s.id=r.student_id
               WHERE r.id=?""", (rid,), one=True)
    if not row:
        flash('Result not found')
        return redirect(url_for('admin_results'))
    return build_result_pdf_response(row)

init_db()

if __name__ == '__main__':
    app.run(debug=True)
